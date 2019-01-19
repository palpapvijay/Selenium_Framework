import openpyxl

# Data Driven - read data from lists.
def dataGenerator():
    li = [['uname1','pass1'],['uname2','pass2'],['uname3','pass3']]
    return li

# Data Driven - read data from Excel
# def dataGenerator():
#     vk = openpyxl.load_workbook("excel path")
#     sh = vk['Sheet1']
#     r = sh.max_row()
#     li=[]
#     li1=[]
#     for i in range(1,r+1):
#         li1=[]
#         un = sh.cell(i,1)
#         up = sh.cell(i,2)
#         li1.insert(0, un.value)
#         li1.insert(1,up.value)
#         li.insert(i-1,li1)
#
#     return li